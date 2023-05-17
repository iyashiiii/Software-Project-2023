# ESSENTIAL FILES ~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
    #1. main.py    -- main program
    #2. admin.txt  -- admin username and password
    #3. excel file -- for overall database

# IMPORTS ~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
from tkinter import *
import tkinter as tk
from tkinter import messagebox
from tkinter import filedialog
from tkinter.filedialog import askopenfile
from PIL import Image, ImageTk
from openpyxl import load_workbook

# ADMIN ACCESS ~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
with open(r"C:\Users\kayth\Desktop\admin.txt") as c:
    l = c.read().split('\n')
    
    
# EXCEL FILE FOR DATA ~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
path = r'C:\Users\kayth\Desktop\data.xlsx'
wb = load_workbook(filename=path, read_only=False)
ws = wb.active
  
# INITIAL WINDOW ~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
class app():
    def __init__(s): # s for self -shortcut purposes
        
        s.root = tk.Tk()

        s.root.geometry('330x190')
        s.root.title('DARK STIMULI CLOTHING')

        s.label = tk.Label(s.root, text='Log-in', font=('Arial',12))
        s.label.pack(padx=20,pady=20)

        s.label1 = tk.Label(s.root, text='Username:', font=('Arial', 12))
        s.label1.place(x=30, y=70)

        s.label2 = tk.Label(s.root, text='Password:', font=('Arial', 12))
        s.label2.place(x=30, y=110)

        s.entry1 = tk.Entry(s.root, bg='#DADEDF')
        s.entry1.place(x=120, y=70)     
        
        s.entry2 = tk.Entry(s.root, bg='#DADEDF')
        s.entry2.place(x=120, y=110)
        
        s.btn = tk.Button(s.root, text='Login', command = s.check)
        s.btn.place(x=160, y=150)
        
        
        
        s.root.mainloop()

# FUNCTIONS ~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~

# check if user is an admin ---
    def check(s):
        s.txt = s.entry1.get()
        s.txt2 = s.entry2.get()
        if s.txt in l and s.txt2 in l:
            s.new_window()
        else:
            messagebox.showerror(title='Message', message='Invalid username and password')
            
# upload function for product image --
    def upload(s):
        s.f_types = [('JPG','*.jpg'),('PNG','*.png'),('JPEG','*.jpeg')]
        s.filename = filedialog.askopenfilename(filetypes=s.f_types)
        s.img = ImageTk.PhotoImage(file=s.filename)
        
        s.frameupload = tk.Frame(s.win)
        s.frameupload.columnconfigure(0, weight=1)
        s.frameupload.columnconfigure(1, weight=1)
        s.frameupload.columnconfigure(2, weight=1)
        
        s.picture = tk.Label(s.frameupload, image = s.img)
        s.picture.grid(row=0, column=0)
        
        s.frameupload.pack()

# place data in excel from New Product page -- 
    def data(s):
        s.product_name = s.pn1.get()
        s.product_category = s.pc1.get()
        s.product_desc = s.pd1.get('1.0',tk.END)
        s.product_price = s.price1.get()
        s.variation = s.colorvar.get('1.0',tk.END)
        
        s.l0 = []
        
        if s.checkbox.get() == 1:
            s.l0.append('/')
        else:
            s.l0.append('')
        if s.checkbox2.get() == 1:
            s.l0.append('/')
        else:
            s.l0.append('')
        if s.checkbox3.get() == 1:
            s.l0.append('/')
        else:
            s.l0.append('')
        if s.checkbox4.get() == 1:
            s.l0.append('/')
        else:
            s.l0.append('')
        if s.checkbox5.get() == 1:
            s.l0.append('/')
        else:
            s.l0.append('')
        if s.checkbox6.get() == 1:
            s.l0.append('/')
        else:
            s.l0.append('')
        
        s.i = ws.max_row
        s.i+=1
        s.entered_data = [s.i, s.product_name, s.product_category, s.product_desc, s.product_price, s.variation]
        s.final = s.entered_data + s.l0
        print(s.final)
        ws.append(s.final)
        wb.save(path)
        messagebox.showinfo(title='Message', message='Data saved')

# place sales data in excel from Enter Sales page -- 
    def data_label(s):
        s.product_sel = s.menu.get()
        s.product_sales = s.sales.get()
        
        for row in range(1, s.fws.max_row):
            for col in s.fws.iter_cols(2,2):
                if col[row].value ==  s.product_sel:
                    #print('Hello world')
                    messagebox.showinfo(title='Message', message='Data saved')
        
        
        
# SECONDARY WINDOWS ~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~

# Main Page selection for all function --
    def new_window(s):
        s.main = tk.Toplevel(s.root)
        s.main.title('Main Page')
        s.main.geometry('500x400')
        s.newlabel = tk.Label(s.main, text='What would you like to do?', font=('Arial',20), height=5)
        s.newlabel.pack(fill='x')#(x=90, y=30)
        
        s.frame = tk.Frame(s.main)
        s.frame.columnconfigure(0, weight=1)
        
        s.btn1 = tk.Button(s.frame, text='New product', font=('Arial', 18), background='#FEFEBE', command = s.window1)
        s.btn1.grid(row=0, column=0, sticky='news')
        
        s.btn2 = tk.Button(s.frame, text='Enter sales', font=('Arial', 18), background='#C7F6B6', command = s.window2)
        s.btn2.grid(row=1, column=0, sticky='news')
        
        s.frame.pack(fill='x')
        
        s.close = tk.Button(s.main, text='Close', font=('Arial', 16), width=8, background='#FFCBCB')
        s.close.pack(side='bottom', pady=30)

        s.main.mainloop()
        
        
# New Product page --       
    def window1(s):
        s.win = tk.Toplevel(s.main)
        s.win.title('New product')
        s.win.geometry('500x800')
        s.pn = tk.Label(s.win, text='Product Name:', font=('Arial',12))
        s.pn.pack(padx=10,pady=2, anchor='nw')
        
        s.pn1 = tk.Entry(s.win, width=50, bg='#DADEDF')
        s.pn1.place(x=120,y=5)
        
        s.pc = tk.Label(s.win, text='Product Category:', font=('Arial',12))
        s.pc.pack(padx=10,pady=2, anchor='nw')
        
        s.pc1 = tk.Entry(s.win, width=50, bg='#DADEDF')
        s.pc1.place(x=145,y=35)
        
        s.pd = tk.Label(s.win, text='Short Product Description:', font=('Arial',12))
        s.pd.pack(padx=10,pady=2, anchor='nw')

        s.pd1 = tk.Text(s.win, height=3, background='#DADEDF')
        s.pd1.pack(padx=10,pady=2)
        
        s.price = tk.Label(s.win, text='Product Price: (number only)', font=('Arial',12))
        s.price.pack(padx=10,pady=2, anchor='nw')
        
        s.price1 = tk.Entry(s.win, width=30, bg='#DADEDF')
        s.price1.place(x=218,y=146)
        
        s.cl = tk.Label(s.win, text='Available Sizes:', font=('Arial', 12))
        s.cl.pack(padx=10,pady=2, anchor='nw')
        
        s.checkbox = tk.IntVar()
        s.checkbox2 = tk.IntVar()
        s.checkbox3 = tk.IntVar()
        s.checkbox4 = tk.IntVar()
        s.checkbox5 = tk.IntVar()
        s.checkbox6 = tk.IntVar()
        
        s.c = tk.Checkbutton(s.win, text='XS', font=('Arial', 8), variable=s.checkbox, bg='#DADEDF')
        s.c.pack(padx=10,pady=2, anchor='nw')
        
        s.c2 = tk.Checkbutton(s.win, text='S', font=('Arial', 8), variable=s.checkbox2, bg='#DADEDF')
        s.c2.pack(padx=10,pady=2, anchor='nw')
        
        s.c3 = tk.Checkbutton(s.win, text='M', font=('Arial', 8), variable=s.checkbox3, bg='#DADEDF')
        s.c3.pack(padx=10,pady=2, anchor='nw')
        
        s.c4 = tk.Checkbutton(s.win, text='L', font=('Arial', 8), variable=s.checkbox4, bg='#DADEDF')
        s.c4.pack(padx=10,pady=2, anchor='nw')
        
        s.c5 = tk.Checkbutton(s.win, text='XL', font=('Arial', 8), variable=s.checkbox5, bg='#DADEDF')
        s.c5.pack(padx=10,pady=2, anchor='nw')
        
        s.c6 = tk.Checkbutton(s.win, text='XXL', font=('Arial', 8), variable=s.checkbox6, bg='#DADEDF')
        s.c6.pack(padx=10,pady=2, anchor='nw')
    
        s.colors = tk.Label(s.win, text='Available Colors: (separate with comma only, no spaces)', font=('Arial',12))
        s.colors.pack(padx=10,pady=2, anchor='nw')

        s.colorvar = tk.Text(s.win, height=3, background='#DADEDF')
        s.colorvar.pack(padx=10,pady=2)
        
        s.upl = tk.Button(s.win, text = 'Upload Product Image & Preview', font=('Arial', 10), background='#DADEDF', command = s.upload)
        s.upl.pack(padx=10,pady=5,anchor='nw')
        
        s.submit = tk.Button(s.win, text='Submit', font=('Arial', 16), width=8, background='#FFCBCB', command = s.data)
        s.submit.pack(side='bottom', pady=20)
        
        
        
        s.win.mainloop()

# Enter Sales page --         
    def window2(s):
        s.win2 = tk.Toplevel(s.main)
        s.win2.title('Enter sales')
        s.win2.geometry('600x500')
        
        s.label2 = tk.Label(s.win2, text='Select a product and enter sales for today', font=('Arial',20), height=5)
        s.label2.pack(fill='x')#(x=90, y=30)
        
        
        s.fpath = r'C:\Users\21-258831\Desktop\data.xlsx'
        s.fwb = load_workbook(filename=path, read_only=False)
        s.fws = s.fwb.active

        s.li = []
        for row in range(1, s.fws.max_row):
            for col in s.fws.iter_cols(2,2):
                s.li.append(col[row].value)
        
       
        s.menu = StringVar()
        s.menu.set('Select a product')
        s.drop = OptionMenu(s.win2, s.menu, *s.li)
        s.drop.config(width=15)
        s.drop.config(height=2)
        s.drop.config(bg='light pink')
        s.drop.config(font=('Arial',12))
        
        s.drop.pack()
        
        s.sales = tk.Entry(s.win2, width=30, bg='#DADEDF')
        s.sales.pack(pady = 8)
        
        s.submitsales = tk.Button(s.win2, text='Submit', font=('Arial', 12), width=8, background='#FFCBCB', command = s.data_label)
        s.submitsales.pack(side='bottom', pady=100)
        
        s.win2.mainloop()
        
        
        
        # Iterate the loop to read the cell values
        
        
    
        
        
    
            

app()
